#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Run designated-target selection, radar coding, and optional ALE in cascade."""

import datetime as _datetime
import os
from pathlib import Path
import pprint
import subprocess
import sys

import openpyxl


# AML
VALID_STEPS = ("selection", "rc", "ale")


def parse_parms(parms_file):
    print("looking for a workflow .parms file at: " + parms_file)
    with open(parms_file, "r") as inp:
        try:
            return eval(inp.read())
        except Exception:
            print("Something wrong with workflow parameters file.")
            raise


def ensure_trailing_sep(path_value):
    path_value = str(path_value)
    if not path_value.endswith(os.sep):
        path_value += os.sep
    return path_value


def resolve_station_ids_to_selection_rows(target_db, exclude_target_ids):
    """Translate user-facing station IDs to selection.py row indexes."""
    if not exclude_target_ids:
        return []

    wanted = {str(target_id).strip() for target_id in exclude_target_ids}
    found = {}

    cell_row_start = 5
    cell_col_id = 4
    cell_col_north = 15

    workbook = openpyxl.load_workbook(target_db, read_only=True, data_only=True)
    sheet = workbook["Database"]

    cell_row = cell_row_start
    content = sheet.cell(row=cell_row, column=cell_col_north).value
    while content is not None:
        target_id = str(sheet.cell(row=cell_row, column=cell_col_id).value).strip()
        if target_id in wanted:
            found[target_id] = cell_row - cell_row_start + 1
        cell_row += 1
        content = sheet.cell(row=cell_row, column=cell_col_north).value

    missing = sorted(wanted - set(found))
    if missing:
        raise ValueError(
            "excludeTargetIds contains station ID(s) not found in targetDB: "
            + ", ".join(missing)
        )

    return [found[target_id] for target_id in exclude_target_ids]


def build_step_parms(workflow):
    project = workflow["project"]
    paths = workflow["paths"]
    selection = workflow.get("selection", {})
    processing = workflow.get("processing", {})
    ale = workflow.get("ale", {})

    out_dir = ensure_trailing_sep(paths["outDir"])
    ale_out_dir = ensure_trailing_sep(paths.get("aleOutDir", out_dir + "ALE" + os.sep))
    station_log = paths.get("stationLog", out_dir + "reflectors.json")

    if "exclTarget" in selection:
        excl_target = selection["exclTarget"]
    else:
        excl_target = resolve_station_ids_to_selection_rows(
            paths["targetDB"],
            selection.get("excludeTargetIds", []),
        )

    selection_parms = {
        "project": project,
        "aoiDir": paths["aoiDir"],
        "targetDB": paths["targetDB"],
        "outDir": out_dir,
        "convFlag": selection.get("convFlag", 1),
        "mapFlag": selection.get("mapFlag", 1),
        "exclTarget": excl_target,
    }

    rc_parms = {
        "project": project,
        "stationLog": station_log,
        "stackLog": paths["stackLog"],
        "outDir": out_dir,
        "precisePosFlag": processing.get("precisePosFlag", 1),
        "plotFlag": processing.get("plotFlag", 1),
        "fullStack": processing.get("fullStack", 1),
        "cropFlag": processing.get("cropFlag", 1),
        "ovsFactor": processing.get("ovsFactor", 1),
    }

    ale_parms = {
        "project": project,
        "stationLog": station_log,
        "stackLog": paths["stackLog"],
        "outDir": ale_out_dir,
        "precisePosFlag": processing.get("precisePosFlag", 1),
        "plotFlag": processing.get("plotFlag", 1),
        "atmoFlag": ale.get("atmoFlag", 0),
        "ovsFactor": processing.get("ovsFactor", 1),
        "cropFlag": processing.get("cropFlag", 1),
    }
    return {
        "selection": selection_parms,
        "rc": rc_parms,
        "ale": ale_parms,
    }


def write_parms(path, parms):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as file:
        file.write(pprint.pformat(parms, sort_dicts=False))
        file.write("\n")


def run_step(step_name, script_name, parms_path, log_file):
    cmd = [sys.executable, "-u", script_name, str(parms_path)]
    header = (
        "\n"
        + "=" * 80
        + f"\nRunning step: {step_name}\n"
        + "Command: "
        + " ".join(cmd)
        + "\n"
        + "=" * 80
        + "\n"
    )
    print(header, end="")
    log_file.write(header)
    log_file.flush()

    process = subprocess.Popen(
        cmd,
        cwd=Path(__file__).resolve().parent,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        bufsize=1,
    )

    for line in process.stdout:
        print(line, end="")
        log_file.write(line)
    process.wait()
    log_file.flush()

    if process.returncode != 0:
        raise RuntimeError(
            f"Step '{step_name}' failed with exit code {process.returncode}. "
            f"See log: {log_file.name}"
        )


def main(workflow):
    steps = workflow.get("steps", ["selection", "rc", "ale"])
    unknown_steps = [step for step in steps if step not in VALID_STEPS]
    if unknown_steps:
        raise ValueError("Unknown workflow step(s): " + ", ".join(unknown_steps))

    step_parms = build_step_parms(workflow)

    out_dir = Path(step_parms["rc"]["outDir"])
    timestamp = _datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    workflow_dir = out_dir / "workflow_parms" / timestamp
    log_path = out_dir / "workflow_logs" / f"cascade_{timestamp}.log"
    log_path.parent.mkdir(parents=True, exist_ok=True)

    generated_parms = {}
    for step, parms in step_parms.items():
        parms_path = workflow_dir / f"{step}.parms"
        write_parms(parms_path, parms)
        generated_parms[step] = parms_path

    scripts = {
        "selection": "selection.py",
        "rc": "mainRC.py",
        "ale": "mainALE.py",
    }

    with open(log_path, "w") as log_file:
        print(f"Workflow log: {log_path}")
        log_file.write(f"Workflow log: {log_path}\n")
        print(f"Generated step parameter files: {workflow_dir}")
        log_file.write(f"Generated step parameter files: {workflow_dir}\n")

        for step in steps:
            if step in ("rc", "ale"):
                station_log = Path(step_parms[step]["stationLog"])
                if not station_log.exists():
                    raise FileNotFoundError(
                        f"{step} requires stationLog, but it does not exist: {station_log}"
                    )
            run_step(step, scripts[step], generated_parms[step], log_file)

        done = "\nCascade workflow completed successfully.\n"
        print(done, end="")
        log_file.write(done)

    print(f"Workflow log written to: {log_path}")
    print(f"Generated step parameter files written to: {workflow_dir}")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        main(parse_parms(sys.argv[1]))
    else:
        raise SystemExit("Usage: python runRadarCodingWorkflow.py <workflow.parms>")
# AML END
